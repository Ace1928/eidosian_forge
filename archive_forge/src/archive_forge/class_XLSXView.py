from __future__ import absolute_import, print_function, division
import itertools
from petl.compat import PY3, text_type
from petl.util.base import Table, data
from petl.io.sources import read_source_from_arg, write_source_from_arg
class XLSXView(Table):

    def __init__(self, filename, sheet=None, range_string=None, min_row=None, min_col=None, max_row=None, max_col=None, read_only=False, **kwargs):
        self.filename = filename
        self.sheet = sheet
        self.range_string = range_string
        self.min_row = min_row
        self.min_col = min_col
        self.max_row = max_row
        self.max_col = max_col
        self.read_only = read_only
        self.kwargs = kwargs

    def __iter__(self):
        import openpyxl
        source = read_source_from_arg(self.filename)
        with source.open('rb') as source2:
            wb = openpyxl.load_workbook(filename=source2, read_only=self.read_only, **self.kwargs)
            if self.sheet is None:
                ws = wb[wb.sheetnames[0]]
            elif isinstance(self.sheet, int):
                ws = wb[wb.sheetnames[self.sheet]]
            else:
                ws = wb[str(self.sheet)]
            if self.range_string is not None:
                rows = ws[self.range_string]
            else:
                rows = ws.iter_rows(min_row=self.min_row, min_col=self.min_col, max_row=self.max_row, max_col=self.max_col)
            for row in rows:
                yield tuple((cell.value for cell in row))
            try:
                wb._archive.close()
            except AttributeError:
                pass