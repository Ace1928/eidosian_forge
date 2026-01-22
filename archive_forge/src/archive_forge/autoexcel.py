from openpyxl import load_workbook

# Load the Excel workbook
workbook = load_workbook("data.xlsx")
sheet = workbook.active

# Update cell values based on conditions
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[1] > 100:
        sheet.cell(row=row[0], column=3).value = "High"
    else:
        sheet.cell(row=row[0], column=3).value = "Low"

# Save the updated workbook
workbook.save("updated_data.xlsx")
