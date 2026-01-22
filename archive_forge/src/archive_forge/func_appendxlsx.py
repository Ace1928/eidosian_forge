from __future__ import absolute_import, print_function, division
import itertools
from petl.compat import PY3, text_type
from petl.util.base import Table, data
from petl.io.sources import read_source_from_arg, write_source_from_arg
def appendxlsx(tbl, filename, sheet=None, write_header=False):
    """
    Appends rows to an existing Excel .xlsx file.
    """
    import openpyxl
    source = read_source_from_arg(filename)
    with source.open('rb') as source2:
        wb = openpyxl.load_workbook(filename=source2, read_only=False)
        if sheet is None:
            ws = wb[wb.sheetnames[0]]
        elif isinstance(sheet, int):
            ws = wb[wb.sheetnames[sheet]]
        else:
            ws = wb[str(sheet)]
        if write_header:
            it = iter(tbl)
            try:
                hdr = next(it)
                flds = list(map(text_type, hdr))
                rows = itertools.chain([flds], it)
            except StopIteration:
                rows = it
        else:
            rows = data(tbl)
        for row in rows:
            ws.append(row)
        target = write_source_from_arg(filename)
        with target.open('wb') as target2:
            wb.save(target2)