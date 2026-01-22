from __future__ import annotations
import mmap
from typing import (
import numpy as np
from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc
from pandas.core.shared_docs import _shared_docs
from pandas.io.excel._base import (
from pandas.io.excel._util import (
class OpenpyxlReader(BaseExcelReader['Workbook']):

    @doc(storage_options=_shared_docs['storage_options'])
    def __init__(self, filepath_or_buffer: FilePath | ReadBuffer[bytes], storage_options: StorageOptions | None=None, engine_kwargs: dict | None=None) -> None:
        """
        Reader using openpyxl engine.

        Parameters
        ----------
        filepath_or_buffer : str, path object or Workbook
            Object to be parsed.
        {storage_options}
        engine_kwargs : dict, optional
            Arbitrary keyword arguments passed to excel engine.
        """
        import_optional_dependency('openpyxl')
        super().__init__(filepath_or_buffer, storage_options=storage_options, engine_kwargs=engine_kwargs)

    @property
    def _workbook_class(self) -> type[Workbook]:
        from openpyxl import Workbook
        return Workbook

    def load_workbook(self, filepath_or_buffer: FilePath | ReadBuffer[bytes], engine_kwargs) -> Workbook:
        from openpyxl import load_workbook
        default_kwargs = {'read_only': True, 'data_only': True, 'keep_links': False}
        return load_workbook(filepath_or_buffer, **default_kwargs | engine_kwargs)

    @property
    def sheet_names(self) -> list[str]:
        return [sheet.title for sheet in self.book.worksheets]

    def get_sheet_by_name(self, name: str):
        self.raise_if_bad_sheet_by_name(name)
        return self.book[name]

    def get_sheet_by_index(self, index: int):
        self.raise_if_bad_sheet_by_index(index)
        return self.book.worksheets[index]

    def _convert_cell(self, cell) -> Scalar:
        from openpyxl.cell.cell import TYPE_ERROR, TYPE_NUMERIC
        if cell.value is None:
            return ''
        elif cell.data_type == TYPE_ERROR:
            return np.nan
        elif cell.data_type == TYPE_NUMERIC:
            val = int(cell.value)
            if val == cell.value:
                return val
            return float(cell.value)
        return cell.value

    def get_sheet_data(self, sheet, file_rows_needed: int | None=None) -> list[list[Scalar]]:
        if self.book.read_only:
            sheet.reset_dimensions()
        data: list[list[Scalar]] = []
        last_row_with_data = -1
        for row_number, row in enumerate(sheet.rows):
            converted_row = [self._convert_cell(cell) for cell in row]
            while converted_row and converted_row[-1] == '':
                converted_row.pop()
            if converted_row:
                last_row_with_data = row_number
            data.append(converted_row)
            if file_rows_needed is not None and len(data) >= file_rows_needed:
                break
        data = data[:last_row_with_data + 1]
        if len(data) > 0:
            max_width = max((len(data_row) for data_row in data))
            if min((len(data_row) for data_row in data)) < max_width:
                empty_cell: list[Scalar] = ['']
                data = [data_row + (max_width - len(data_row)) * empty_cell for data_row in data]
        return data