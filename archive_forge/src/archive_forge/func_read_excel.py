import os
import json
import csv
import xml.etree.ElementTree as ET
from typing import Any, Dict, Union, Optional, Tuple, List
import pandas as pd
import logging
import yaml
import pickle
import configparser
import markdown
import openpyxl
import sqlite3
import PyPDF2
import PIL.Image
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Label, Toplevel
from PIL import Image, ImageTk
import os
import logging
import json
import pandas as pd
def read_excel(self) -> openpyxl.workbook.workbook.Workbook:
    """
        Read an Excel file and return its content as an openpyxl Workbook object, ensuring comprehensive extraction
        of all data types including text, formulas, images, charts, and hyperlinks. This method is designed to handle
        complex Excel files with multiple sheets, providing a robust and complete representation of the workbook.

        :return: The content of the Excel file, including all sheets and their respective elements.
        :rtype: openpyxl.workbook.workbook.Workbook
        """
    try:
        workbook = openpyxl.load_workbook(self.file_path, data_only=False)
        logging.debug(f'Excel data read successfully from {self.file_path}')
        sheets = workbook.sheetnames
        logging.info(f'Workbook contains {len(sheets)} sheets: {sheets}')
        for sheet_name in sheets:
            sheet = workbook[sheet_name]
            num_rows = sheet.max_row
            num_cols = sheet.max_column
            logging.info(f"Sheet '{sheet_name}' contains {num_rows} rows and {num_cols} columns.")
            images = sheet._images
            charts = sheet.charts
            hyperlinks = [cell.hyperlink for row in sheet.iter_rows() for cell in row if cell.hyperlink]
            logging.info(f"Sheet '{sheet_name}' contains {len(images)} images, {len(charts)} charts, and {len(hyperlinks)} hyperlinks.")
        return workbook
    except Exception as e:
        logging.error(f'An error occurred while reading the Excel file at {self.file_path}: {str(e)}')
        raise Exception(f'An error occurred while reading the Excel file: {str(e)}')